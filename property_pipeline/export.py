"""Export functions to produce XLSX/CSV files compatible with 3.0 MonthlySummary."""

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def build_output_dataframe(transactions: list[dict], labels: list[dict]) -> pd.DataFrame:
    """Join canonical transactions with labels into the output schema.

    Output columns (matching 3.0 MonthlySummary expectations):
        Date (index), Account, Amount, Subcategory, Memo, Property, Description, Cat, Subcat
    """
    tx_df = pd.DataFrame(transactions)
    lab_df = pd.DataFrame(labels)

    tx_df = tx_df[tx_df["is_superseded"] == 0].copy()

    merged = tx_df.merge(lab_df, on="tx_id", how="left")

    # posted_date should be YYYY-MM-DD; coerce and handle bad/mixed data
    date_ser = merged["posted_date"].astype(str)
    merged["Date"] = pd.to_datetime(date_ser, format="%Y-%m-%d", errors="coerce")
    bad = merged["Date"].isna() & (date_ser != "").values
    if bad.any():
        merged.loc[bad, "Date"] = pd.to_datetime(date_ser[bad], dayfirst=True, errors="coerce")
    # Drop rows where date could not be parsed (e.g. wrong column data)
    merged = merged.dropna(subset=["Date"])
    merged["Account"] = merged["source_account"]
    merged["Amount"] = merged["amount"]
    merged["Subcategory"] = merged["effective_subcategory"]
    merged["Memo"] = merged["memo"]
    merged["Property"] = merged["property_code"].fillna("")
    desc_y = merged.get("description_y")
    desc_x = merged.get("description_x")
    if desc_y is not None:
        merged["Description"] = desc_y.fillna("").astype(str)
    elif desc_x is not None:
        merged["Description"] = desc_x.fillna("").astype(str)
    else:
        merged["Description"] = ""
    merged["Cat"] = merged["category"].fillna("")
    merged["Subcat"] = merged["subcategory"].fillna("")

    merged.set_index("Date", inplace=True)
    merged.sort_index(inplace=True)

    output_cols = ["Account", "Amount", "Subcategory", "Memo", "Property", "Description", "Cat", "Subcat"]
    result = merged[output_cols].copy()

    return result


# Number of rows in Lists sheet; blank rows allow users to add new property codes/categories/subcategories
VALIDATION_LIST_ROWS = 250


def _set_column_widths(
    ws,
    skip_col_letters: set[str] | None = None,
    width_scale: dict[str, float] | None = None,
) -> None:
    """Set each column width to the max character length in that column.
    skip_col_letters: do not set width for these columns (e.g. tx_id).
    width_scale: optional dict col_letter -> scale (e.g. {'E': 0.5} for 50% width for Memo).
    """
    skip_col_letters = skip_col_letters or set()
    width_scale = width_scale or {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in skip_col_letters:
            continue
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        if max_len > 0:
            w = max_len
            if col_letter in width_scale:
                w = max(1, round(w * width_scale[col_letter]))
            ws.column_dimensions[col_letter].width = min(w, 255)


def write_xlsx(
    df: pd.DataFrame,
    output_path: Path,
    property_codes: list[str] | None = None,
    categories: list[str] | None = None,
    subcategories: list[str] | None = None,
) -> None:
    """Write the dataframe to XLSX.

    If property_codes, categories, or subcategories are provided, adds a 'Lists' sheet
    and sets Excel data validation (dropdowns) on the Data sheet for Property, Cat, and Subcat.
    The Lists sheet has room for extra rows so users can add new values that then appear in the dropdowns.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    use_validation = property_codes is not None or categories is not None or subcategories is not None

    if not use_validation:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Data", index=True)
            ws_data = writer.sheets["Data"]
            _set_column_widths(ws_data, width_scale={"E": 0.5})
            _set_auto_filter(ws_data)
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=True)
        wb = writer.book
        ws_data = wb["Data"]
        # Insert Lists after Data so Data remains the first sheet
        ws_lists = wb.create_sheet("Lists", 1)
        ws_lists["A1"] = "Property"
        ws_lists["B1"] = "Category"
        ws_lists["C1"] = "Subcategory"
        for i in range(min(len(property_codes or []), VALIDATION_LIST_ROWS)):
            ws_lists.cell(row=i + 2, column=1, value=(property_codes or [])[i])
        for i in range(min(len(categories or []), VALIDATION_LIST_ROWS)):
            ws_lists.cell(row=i + 2, column=2, value=(categories or [])[i])
        for i in range(min(len(subcategories or []), VALIDATION_LIST_ROWS)):
            ws_lists.cell(row=i + 2, column=3, value=(subcategories or [])[i])

        ws_data = wb["Data"]
        max_row = max(ws_data.max_row or 2, 2)
        range_end = VALIDATION_LIST_ROWS + 1
        dv_prop = DataValidation(
            type="list",
            formula1=f"Lists!$A$2:$A${range_end}",
            allow_blank=True,
        )
        dv_cat = DataValidation(
            type="list",
            formula1=f"Lists!$B$2:$B${range_end}",
            allow_blank=True,
        )
        dv_sub = DataValidation(
            type="list",
            formula1=f"Lists!$C$2:$C${range_end}",
            allow_blank=True,
        )
        dv_prop.add(f"F2:F{max_row}")
        dv_cat.add(f"H2:H{max_row}")
        dv_sub.add(f"I2:I{max_row}")
        ws_data.add_data_validation(dv_prop)
        ws_data.add_data_validation(dv_cat)
        ws_data.add_data_validation(dv_sub)
        _set_column_widths(ws_data, width_scale={"E": 0.5})  # Memo at 50%
        _set_auto_filter(ws_data)


def _set_auto_filter(ws) -> None:
    """Turn on Excel auto-filter for the used range (all column headers)."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def write_csv(df: pd.DataFrame, output_path: Path) -> None:
    """Write the dataframe to CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path)


def write_review_queue(
    transactions: list[dict],
    labels: list[dict],
    output_path: Path,
    property_codes: list[str] | None = None,
    categories: list[str] | None = None,
    subcategories: list[str] | None = None,
) -> int:
    """Write review queue XLSX with only needs_review=1 rows.

    If property_codes, categories, or subcategories are provided, adds a 'Lists' sheet
    and dropdown validation on property_code, category, subcategory.
    Returns the count of review items.
    """
    tx_df = pd.DataFrame(transactions)
    lab_df = pd.DataFrame(labels)
    tx_df = tx_df[tx_df["is_superseded"] == 0].copy()

    merged = tx_df.merge(lab_df, on="tx_id", how="left")
    review = merged[merged["needs_review"] == 1].copy()

    if review.empty:
        return 0

    review["Date"] = pd.to_datetime(review["posted_date"])
    review.set_index("Date", inplace=True)
    review.sort_index(inplace=True)

    cols = [
        "tx_id", "source_bank", "source_account", "amount",
        "counterparty", "reference", "memo", "type",
        "effective_subcategory",
        "property_code", "category", "subcategory",
        "confidence", "rule_id", "rule_strength",
    ]
    existing = [c for c in cols if c in review.columns]
    out = review[existing]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    use_validation = property_codes is not None or categories is not None or subcategories is not None

    skip_tx_id = set()
    memo_scale = {}
    if "tx_id" in existing:
        skip_tx_id.add(get_column_letter(2 + existing.index("tx_id")))
    if "memo" in existing:
        memo_scale[get_column_letter(2 + existing.index("memo"))] = 0.5

    try:
        writer_ctx = pd.ExcelWriter(output_path, engine="openpyxl")
    except PermissionError as e:
        raise PermissionError(
            f"Cannot write to {output_path}. Close the file if it is open in Excel or another program, then try again."
        ) from e

    with writer_ctx as writer:
        out.to_excel(writer, sheet_name="Review", index=True)
        wb = writer.book
        ws = wb["Review"]
        _set_column_widths(ws, skip_col_letters=skip_tx_id, width_scale=memo_scale)
        _set_auto_filter(ws)

        if use_validation:
            # Insert Lists after Review so Review remains the first sheet
            ws_lists = wb.create_sheet("Lists", 1)
            ws_lists["A1"] = "Property"
            ws_lists["B1"] = "Category"
            ws_lists["C1"] = "Subcategory"
            for i in range(min(len(property_codes or []), VALIDATION_LIST_ROWS)):
                ws_lists.cell(row=i + 2, column=1, value=(property_codes or [])[i])
            for i in range(min(len(categories or []), VALIDATION_LIST_ROWS)):
                ws_lists.cell(row=i + 2, column=2, value=(categories or [])[i])
            for i in range(min(len(subcategories or []), VALIDATION_LIST_ROWS)):
                ws_lists.cell(row=i + 2, column=3, value=(subcategories or [])[i])

            ws = wb["Review"]
            max_row = max(ws.max_row or 2, 2)
            range_end = VALIDATION_LIST_ROWS + 1
            try:
                i_prop = existing.index("property_code")
                i_cat = existing.index("category")
                i_sub = existing.index("subcategory")
            except ValueError:
                i_prop = i_cat = i_sub = None
            if i_prop is not None:
                col_prop = get_column_letter(2 + i_prop)
                dv_prop = DataValidation(
                    type="list",
                    formula1=f"Lists!$A$2:$A${range_end}",
                    allow_blank=True,
                )
                dv_prop.add(f"{col_prop}2:{col_prop}{max_row}")
                ws.add_data_validation(dv_prop)
            if i_cat is not None:
                col_cat = get_column_letter(2 + i_cat)
                dv_cat = DataValidation(
                    type="list",
                    formula1=f"Lists!$B$2:$B${range_end}",
                    allow_blank=True,
                )
                dv_cat.add(f"{col_cat}2:{col_cat}{max_row}")
                ws.add_data_validation(dv_cat)
            if i_sub is not None:
                col_sub = get_column_letter(2 + i_sub)
                dv_sub = DataValidation(
                    type="list",
                    formula1=f"Lists!$C$2:$C${range_end}",
                    allow_blank=True,
                )
                dv_sub.add(f"{col_sub}2:{col_sub}{max_row}")
                ws.add_data_validation(dv_sub)

    return len(out)


def write_diagnostic_ddcheck(
    transactions: list[dict],
    labels: list[dict],
    output_path: Path,
) -> None:
    """Write DDCheck equivalent: all direct debits and Beals for mortgage/property checking."""
    import re

    tx_df = pd.DataFrame(transactions)
    lab_df = pd.DataFrame(labels)
    tx_df = tx_df[tx_df["is_superseded"] == 0].copy()
    merged = tx_df.merge(lab_df, on="tx_id", how="left")

    mask_dd = merged["effective_subcategory"].fillna("").str.contains(
        r"DIRECT[ ]?DEBIT|Direct Debit", case=False, regex=True
    )
    mask_beals = merged["memo"].fillna("").str.match(r"^BEALS.*$", case=False)
    dd = merged[mask_dd | mask_beals].copy()

    dd["Date"] = pd.to_datetime(dd["posted_date"])
    dd.set_index("Date", inplace=True)
    dd.sort_index(inplace=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    dd.to_csv(output_path)


def write_diagnostic_catcheck(
    transactions: list[dict],
    labels: list[dict],
    output_path: Path,
) -> None:
    """Write CatCheck equivalent: all categorised transactions."""
    tx_df = pd.DataFrame(transactions)
    lab_df = pd.DataFrame(labels)
    tx_df = tx_df[tx_df["is_superseded"] == 0].copy()
    merged = tx_df.merge(lab_df, on="tx_id", how="left")

    merged["Date"] = pd.to_datetime(merged["posted_date"])
    merged.set_index("Date", inplace=True)
    merged.sort_index(inplace=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged.to_csv(output_path)
